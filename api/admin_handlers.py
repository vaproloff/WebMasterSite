from cmath import inf

from fastapi import APIRouter, Depends
from fastapi import Request
from fastapi.encoders import jsonable_encoder
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse
from datetime import datetime
from datetime import timedelta
from itertools import groupby

from sqlalchemy.ext.asyncio import AsyncSession

from api.actions.indicators import _get_indicators_from_db, _get_top_query, _get_top_url
from api.actions.query_url_merge import _get_merge_with_pagination, _get_merge_query, _get_merge_with_pagination_sort, \
    _get_merge_with_pagination_and_like, _get_merge_with_pagination_and_like_sort
from api.actions.utils import get_day_of_week
from api.auth.auth_config import current_user, RoleChecker
from api.auth.models import User
from api.config.utils import get_config_names, get_group_names
from db.models import QueryUrlsMergeLogs
from db.session import get_db_general, connect_db
from api.actions.urls import _get_urls_with_pagination
from api.actions.urls import _get_urls_with_pagination_and_like
from api.actions.urls import _get_urls_with_pagination_sort
from api.actions.urls import _get_urls_with_pagination_and_like_sort

from api.actions.queries import _get_urls_with_pagination_query
from api.actions.queries import _get_urls_with_pagination_and_like_query
from api.actions.queries import _get_urls_with_pagination_sort_query
from api.actions.queries import _get_urls_with_pagination_and_like_sort_query

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
import io

import csv

from db.utils import get_all_dates

admin_router = APIRouter()

templates = Jinja2Templates(directory="static")

date_format_2 = "%Y-%m-%d"
date_format_out = "%d.%m.%Y"


def pad_list_with_zeros_excel(lst, amount):
    if len(lst) < amount:
        padding = [0] * (amount - len(lst))
        lst.extend(padding)
    return lst


@admin_router.post("/generate_excel_url/")
async def generate_excel(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    wb = Workbook()
    ws = wb.active
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "Url")
    for i in range(4):
        main_header.insert(1, "Result")
    ws.append(main_header)
    header = ["Position", "Click", "R", "CTR"] * (int(data_request["amount"]) + 4)
    header.insert(0, "")
    ws.append(header)
    start = 0
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("Result")
    main_header = main_header[::-1]
    while True:
        start_el = (start * 50)
        if data_request["sort_result"]:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination_sort(start_el, data_request["length"],
                                                            start_date, end_date,
                                                            data_request["sort_desc"], async_session)
            else:
                urls = await _get_urls_with_pagination_and_like_sort(start_el, data_request["length"],
                                                                     start_date, end_date,
                                                                     data_request["search_text"],
                                                                     data_request["sort_desc"],
                                                                     async_session)
        else:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination(start_el, data_request["length"],
                                                       start_date, end_date, async_session)
            else:
                urls = await _get_urls_with_pagination_and_like(start_el, data_request["length"],
                                                                start_date, end_date,
                                                                data_request["search_text"],
                                                                async_session)
        start += 1
        try:
            if urls:
                urls.sort(key=lambda x: x[-1])
            grouped_data = [(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                            groupby(urls, key=lambda x: x[-1])]
        except TypeError as e:
            break
        if len(grouped_data) == 0:
            break
        for el in grouped_data:
            info = {}
            res = []
            total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
            for k, stat in enumerate(el[1]):
                info[stat[0].strftime(date_format_out)] = [stat[1], stat[2], stat[3], stat[4]]
                total_clicks += stat[2]
                position += stat[1]
                impressions += stat[3]
                ctr += stat[4]
                if stat[1] > 0:
                    count += 1
            if count > 0:
                info["Result"] = [round(position / count, 2), total_clicks, impressions, round(ctr / count, 2)]
            else:
                info["Result"] = [0, total_clicks, impressions, 0]
            res.append(el[0])
            for el in main_header:
                if el in info:
                    res.extend(info[el])
                else:
                    res.extend([0, 0, 0, 0])
            ws.append(res)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(io.BytesIO(output.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment;filename='data.xlsx'"})


@admin_router.post("/generate_csv_urls/")
async def generate_excel(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    ws = []
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "Url")
    ws.append(main_header)
    header = ["Position", "Click", "R", "CTR"] * (int(data_request["amount"]) + 4)
    header.insert(0, "")
    for i in range(4):
        main_header.insert(1, "Result")
    ws.append(header)
    start = 0
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("Result")
    main_header = main_header[::-1]
    while True:
        start_el = (start * 50) + 1
        if data_request["sort_result"]:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination_sort(start_el, data_request["length"],
                                                            start_date, end_date,
                                                            data_request["sort_desc"], async_session)
            else:
                urls = await _get_urls_with_pagination_and_like_sort(start_el, data_request["length"],
                                                                     start_date, end_date,
                                                                     data_request["search_text"],
                                                                     data_request["sort_desc"],
                                                                     async_session)
        else:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination(start_el, data_request["length"],
                                                       start_date, end_date, async_session)
            else:
                urls = await _get_urls_with_pagination_and_like(start_el, data_request["length"],
                                                                start_date, end_date,
                                                                data_request["search_text"],
                                                                async_session)
        start += 1
        try:
            if urls:
                urls.sort(key=lambda x: x[-1])
            grouped_data = [(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                            groupby(urls, key=lambda x: x[-1])]
        except TypeError as e:
            break
        if len(grouped_data) == 0:
            break
        for el in grouped_data:
            info = {}
            res = []
            total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
            for k, stat in enumerate(el[1]):
                info[stat[0].strftime(date_format_out)] = [stat[1], stat[2], stat[3], stat[4]]
                total_clicks += stat[2]
                position += stat[1]
                impressions += stat[3]
                ctr += stat[4]
                if stat[1] > 0:
                    count += 1
            if count > 0:
                info["Result"] = [round(position / count, 2), total_clicks, impressions, round(ctr / count, 2)]
            else:
                info["Result"] = [0, total_clicks, impressions, 0]
            res.append(el[0])
            for el in main_header:
                if el in info:
                    res.extend(info[el])
                else:
                    res.extend([0, 0, 0, 0])
            ws.append(res)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(ws)
    output.seek(0)

    return StreamingResponse(content=output.getvalue(),
                             headers={"Content-Disposition": "attachment;filename='data.csv'"})


@admin_router.post("/generate_excel_queries/")
async def generate_excel(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    wb = Workbook()
    ws = wb.active
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "Url")
    for i in range(4):
        main_header.insert(1, "Result")
    ws.append(main_header)
    header = ["Position", "Click", "R", "CTR"] * (int(data_request["amount"]) + 4)
    header.insert(0, "")
    ws.append(header)
    start = 0
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("Result")
    main_header = main_header[::-1]
    while True:
        start_el = (start * 50) + 1
        if data_request["sort_result"]:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination_sort_query(start_el, data_request["length"], start_date,
                                                                  end_date, data_request["sort_desc"],
                                                                  async_session)
            else:
                urls = await _get_urls_with_pagination_and_like_sort_query(start_el, data_request["length"],
                                                                           start_date, end_date,
                                                                           data_request["search_text"],
                                                                           data_request["sort_desc"],
                                                                           async_session)
        else:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination_query(start_el, data_request["length"], start_date,
                                                             end_date, async_session)
            else:
                urls = await _get_urls_with_pagination_and_like_query(start_el, data_request["length"],
                                                                      start_date, end_date, data_request["search_text"],
                                                                      async_session)
        start += 1
        try:
            if urls:
                urls.sort(key=lambda x: x[-1])
            grouped_data = [(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                            groupby(urls, key=lambda x: x[-1])]
        except TypeError as e:
            break
        if len(grouped_data) == 0:
            break
        for el in grouped_data:
            info = {}
            res = []
            total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
            for k, stat in enumerate(el[1]):
                info[stat[0].strftime(date_format_out)] = [stat[1], stat[2], stat[3], stat[4]]
                total_clicks += stat[2]
                position += stat[1]
                impressions += stat[3]
                ctr += stat[4]
                if stat[1] > 0:
                    count += 1
            if count > 0:
                info["Result"] = [round(position / count, 2), total_clicks, impressions, round(ctr / count, 2)]
            else:
                info["Result"] = [0, total_clicks, impressions, 0]
            res.append(el[0])
            for el in main_header:
                if el in info:
                    res.extend(info[el])
                else:
                    res.extend([0, 0, 0, 0])
            ws.append(res)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(io.BytesIO(output.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment;filename='data.xlsx'"})


@admin_router.post("/generate_csv_queries/")
async def generate_excel(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    ws = []
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "Url")
    for i in range(4):
        main_header.insert(1, "Result")
    ws.append(main_header)
    header = ["Position", "Click", "R", "CTR"] * (int(data_request["amount"]) + 4)
    header.insert(0, "")
    ws.append(header)
    start = 0
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("Result")
    main_header = main_header[::-1]
    while True:
        start_el = (start * 50) + 1
        if data_request["sort_result"]:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination_sort_query(start_el, data_request["length"], start_date,
                                                                  end_date, data_request["sort_desc"],
                                                                  async_session)
            else:
                urls = await _get_urls_with_pagination_and_like_sort_query(start_el, data_request["length"],
                                                                           start_date, end_date,
                                                                           data_request["search_text"],
                                                                           data_request["sort_desc"],
                                                                           async_session)
        else:
            if data_request["search_text"] == "":
                urls = await _get_urls_with_pagination_query(start_el, data_request["length"], start_date,
                                                             end_date, async_session)
            else:
                urls = await _get_urls_with_pagination_and_like_query(start_el, data_request["length"],
                                                                      start_date, end_date, data_request["search_text"],
                                                                      async_session)
        start += 1
        try:
            if urls:
                urls.sort(key=lambda x: x[-1])
            grouped_data = [(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                            groupby(urls, key=lambda x: x[-1])]
        except TypeError as e:
            break
        if len(grouped_data) == 0:
            break
        for el in grouped_data:
            res = []
            info = {}
            total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
            for k, stat in enumerate(el[1]):
                info[stat[0].strftime(date_format_out)] = [stat[1], stat[2], stat[3], stat[4]]
                total_clicks += stat[2]
                position += stat[1]
                impressions += stat[3]
                ctr += stat[4]
                if stat[1] > 0:
                    count += 1
            if count > 0:
                info["Result"] = [round(position / count, 2), total_clicks, impressions, round(ctr / count, 2)]
            else:
                info["Result"] = [0, total_clicks, impressions, 0]
            res.append(el[0])
            for el in main_header:
                if el in info:
                    res.extend(info[el])
                else:
                    res.extend([0, 0, 0, 0])

            ws.append(res)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(ws)
    output.seek(0)

    return StreamingResponse(content=output.getvalue(),
                             headers={"Content-Disposition": "attachment;filename='data.csv'"})


def pad_list_with_zeros(lst, amount):
    if len(lst) < amount:
        padding = [f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #B9BDBC'>
            <span style='font-size: 18px'><span style='color:red'>NAN</span></span><br>
            <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 20px'>CTR <span style='color:red'>NAN%</span></span><br>
            <span style='font-size: 10px'><span style='color:red'>NAN</span></span> <span style='font-size: 10px; margin-left: 30px'>R <span style='color:red'>NAN%</span></span>
            </div>"""] * (amount - len(lst))
        lst.extend(padding)
    return lst


@admin_router.get("/")
async def login_page(request: Request, user: User = Depends(current_user)):
    return templates.TemplateResponse("login.html", {"request": request, "user": user})


@admin_router.get("/register")
async def register(request: Request, user: User = Depends(current_user)):
    return templates.TemplateResponse("register.html", {"request": request, "user": user})


@admin_router.get("/info-urls")
async def get_urls(request: Request,
                   user: User = Depends(current_user),
                   session: AsyncSession = Depends(get_db_general)):
    group_name = request.session["group"].get("name", "")
    config_names = [elem[0] for elem in (await get_config_names(session, user, group_name))]

    group_names = await get_group_names(session, user)

    return templates.TemplateResponse("urls-info.html",
                                      {"request": request,
                                       "user": user,
                                       "config_names": config_names,
                                       "group_names": group_names})


@admin_router.post("/get-urls")
async def get_urls(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    if data_request["sort_result"]:
        if data_request["search_text"] == "":
            urls = await _get_urls_with_pagination_sort(data_request["start"], data_request["length"], start_date,
                                                        end_date, data_request["sort_desc"], async_session)
        else:
            urls = await _get_urls_with_pagination_and_like_sort(data_request["start"], data_request["length"],
                                                                 start_date, end_date, data_request["search_text"],
                                                                 data_request["sort_desc"],
                                                                 async_session)
    else:
        if data_request["search_text"] == "":
            urls = await _get_urls_with_pagination(data_request["start"], data_request["length"], start_date, end_date,
                                                   async_session)
        else:
            urls = await _get_urls_with_pagination_and_like(data_request["start"], data_request["length"], start_date,
                                                            end_date, data_request["search_text"],
                                                            async_session)
    try:
        if urls:
            urls.sort(key=lambda x: x[-1])
        grouped_data = [(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                        groupby(urls, key=lambda x: x[-1])]
    except TypeError as e:
        if urls is None:
            pass
        return JSONResponse({"data": []})
    if len(grouped_data) == 0:
        return JSONResponse({"data": []})
    data = []
    for el in grouped_data:
        res = {"url":
                   f"<div style='width:355px; height: 55px; overflow: auto; white-space: nowrap;'><span>{el[0]}</span></div>"}
        total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
        for k, stat in enumerate(el[1]):
            up = 0
            if k + 1 < len(el[1]):
                up = round(el[1][k][1] - el[1][k - 1][1], 2)
            color = "#9DE8BD"
            color_text = "green"
            if up > 0:
                color = "#FDC4BD"
                color_text = "red"
            if stat[1] <= 3:
                color = "#B4D7ED"
                color_text = "blue"
            res[stat[0].strftime(
                date_format_2)] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: {color}'>
            <span style='font-size: 18px'>{stat[1]}</span><span style="margin-left: 5px; font-size: 10px; color: {color_text}">{abs(up)}</span><br>
            <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 20px'>CTR {stat[4]}%</span><br>
            <span style='font-size: 10px'>{stat[2]}</span> <span style='font-size: 10px; margin-left: 20px'>R {int(stat[3])}</span>
            </div>"""
            total_clicks += stat[2]
            position += stat[1]
            impressions += stat[3]
            ctr += stat[4]
            if stat[1] > 0:
                count += 1
            if k == len(el[1]) - 1:
                res["result"] = res.get("result", "")
                if count > 0:
                    res["result"] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
                              <span style='font-size: 15px'>Позиция:{round(position / count, 2)}</span>
                              <span style='font-size: 15px'>Клики:{total_clicks}</span>
                              <span style='font-size: 9px'>Показы:{impressions}</span>
                              <span style='font-size: 7px'>ctr:{round(ctr / count, 2)}%</span>
                              </div>"""
                else:
                    res["result"] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
                              <span style='font-size: 15px'>Позиция:{0}</span>
                              <span style='font-size: 15px'>Клики:{total_clicks}</span>
                              <span style='font-size: 9px'>Показы:{impressions}</span>
                              <span style='font-size: 9px'>ctr:{0}%</span>
                              </div>"""
        data.append(res)
    json_data = jsonable_encoder(data)

    # return JSONResponse({"data": json_data, "recordsTotal": limit, "recordsFiltered": 50000})
    return JSONResponse({"data": json_data})


@admin_router.get("/info-queries")
async def get_queries(request: Request, user: User = Depends(current_user),
                      session: AsyncSession = Depends(get_db_general)):
    group_name = request.session["group"].get("name", "")
    config_names = [elem[0] for elem in (await get_config_names(session, user, group_name))]

    group_names = await get_group_names(session, user)

    return templates.TemplateResponse("queries-info.html",
                                      {"request": request,
                                       "user": user,
                                       "config_names": config_names,
                                       "group_names": group_names})


@admin_router.post("/get-queries")
async def get_queries(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    if data_request["sort_result"]:
        if data_request["search_text"] == "":
            urls = await _get_urls_with_pagination_sort_query(data_request["start"], data_request["length"], start_date,
                                                              end_date, data_request["sort_desc"],
                                                              async_session)
        else:
            urls = await _get_urls_with_pagination_and_like_sort_query(data_request["start"], data_request["length"],
                                                                       start_date, end_date,
                                                                       data_request["search_text"],
                                                                       data_request["sort_desc"],
                                                                       async_session)
    else:
        if data_request["search_text"] == "":
            urls = await _get_urls_with_pagination_query(data_request["start"], data_request["length"], start_date,
                                                         end_date, async_session)
        else:
            urls = await _get_urls_with_pagination_and_like_query(data_request["start"], data_request["length"],
                                                                  start_date, end_date, data_request["search_text"],
                                                                  async_session)
    try:
        if urls:
            urls.sort(key=lambda x: x[-1])
        grouped_data = [(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                        groupby(urls, key=lambda x: x[-1])]
    except TypeError as e:
        return JSONResponse({"data": []})

    if len(grouped_data) == 0:
        return JSONResponse({"data": []})

    data = []
    for el in grouped_data:
        res = {"query":
                   f"<div style='width:355px; height: 55px; overflow: auto; white-space: nowrap;'><span>{el[0]}</span></div>"}
        total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
        for k, stat in enumerate(el[1]):
            up = 0
            if k + 1 < len(el[1]):
                up = round(el[1][k][1] - el[1][k - 1][1], 2)
            color = "#9DE8BD"
            color_text = "green"
            if up > 0:
                color = "#FDC4BD"
                color_text = "red"
            if stat[1] <= 3:
                color = "#B4D7ED"
                color_text = "blue"
            res[stat[0].strftime(
                date_format_2)] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: {color}'>
              <span style='font-size: 18px'>{stat[1]}</span><span style="margin-left: 5px; font-size: 10px; color: {color_text}">{abs(up)}</span><br>
              <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 10px'>CTR {stat[4]}%</span><br>
              <span style='font-size: 10px'>{stat[2]}</span> <span style='font-size: 10px; margin-left: 20px'>R {int(stat[3])}</span>
              </div>"""
            total_clicks += stat[2]
            position += stat[1]
            impressions += stat[3]
            ctr += stat[4]
            if stat[1] > 0:
                count += 1
            if k == len(el[1]) - 1:
                res["result"] = res.get("result", "")
                if count > 0:
                    res["result"] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
                              <span style='font-size: 15px'>Позиция:{round(position / count, 2)}</span>
                              <span style='font-size: 15px'>Клики:{total_clicks}</span>
                              <span style='font-size: 8px'>Показы:{impressions}</span>
                              <span style='font-size: 7px'>ctr:{round(ctr / count, 2)}%</span>
                              </div>"""
                else:
                    res["result"] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
                              <span style='font-size: 15px'>Позиция:{0}</span>
                              <span style='font-size: 15px'>Клики:{total_clicks}</span>
                              <span style='font-size: 8px'>Показы:{impressions}</span>
                              <span style='font-size: 8px'>ctr:{0}%</span>
                              </div>"""
        data.append(res)

    json_data = jsonable_encoder(data)

    # return JSONResponse({"data": json_data, "recordsTotal": limit, "recordsFiltered": 50000})
    return JSONResponse({"data": json_data})


@admin_router.get("/info-all-history")
async def get_all_history(
        request: Request,
        user: User = Depends(current_user),
        session: AsyncSession = Depends(get_db_general)):
    group_name = request.session["group"].get("name", "")
    config_names = [elem[0] for elem in (await get_config_names(session, user, group_name))]

    group_names = await get_group_names(session, user)

    return templates.TemplateResponse("all-history.html",
                                      {"request": request,
                                       "user": user,
                                       "config_names": config_names,
                                       "group_names": group_names})


@admin_router.post("/get-all-history")
async def post_all_history(
        request: Request, data_request: dict,
        user: User = Depends(current_user)
):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    indicators = await _get_indicators_from_db(start_date,
                                               end_date, async_session)

    if indicators:
        indicators.sort(key=lambda x: x[0])
    try:
        grouped_data = [(key, sorted(list(group), key=lambda x: x[2])) for key, group in
                        groupby(indicators, key=lambda x: x[0])]
    except TypeError as e:
        return JSONResponse({"data": []})

    if len(grouped_data) == 0:
        return JSONResponse({"data": []})

    data = []
    for count, el in enumerate(grouped_data):
        res = {"query":
                   f"<div style='width:355px; height: 55px; overflow: auto; white-space: nowrap;'><span>{el[0]}</span></div>"}
        prev_value = -inf
        value_sum, count = 0, 0
        for name, value, date in el[1]:
            if value >= prev_value:
                color = "#9DE8BD"  # green
            else:
                color = "#FDC4BD"  # red
            if value > 0:
                value_sum += value
                count += 1
                if el[0] != "TOTAL_CTR":
                    res[date.strftime(
                        date_format_2)] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: {color}; text-align: center; display: flex; align-items: center; justify-content: center;'>
                                            <span style='font-size: 18px'>{value}</span>
                                            </div>"""
                else:
                    res[date.strftime(
                        date_format_2)] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: {color}; text-align: center; display: flex; align-items: center; justify-content: center;'>
                                            <span style='font-size: 18px'>{value}%</span>
                                            </div>"""
                prev_value = value

        if el[0] in ("AVG_CLICK_POSITION", "TOTAL_CTR", "AVG_SHOW_POSITION"):
            if count > 0:
                value_sum = round(value_sum / count, 2)

        if el[0] != "TOTAL_CTR":
            res["result"] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD; text-align: center; display: flex; align-items: center; justify-content: center;'>
                                    <span style='font-size: 18px'>{value_sum}</span>
                                    </div>"""
        else:
            res["result"] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD; text-align: center; display: flex; align-items: center; justify-content: center;'>
                                    <span style='font-size: 18px'>{value_sum}%</span>
                                    </div>"""
        data.append(res)
    data[-1], data[-2] = data[-2], data[-1]

    TOP = 3, 5, 10, 20, 30
    query_front = []
    for top in TOP:
        grouped_data_sum = {
            "query":
                f"<div style='width:355px; height: 55px; overflow: auto; white-space: nowrap;'><span>TOP {top}</span></div>"
        }
        query_top = await _get_top_query(start_date, end_date, top, async_session)
        query_top.sort(key=lambda x: x[-1])
        total_position, total_clicks, total_impression, total_count, count_for_avg = 0, 0, 0, 0, 0
        for position, clicks, impression, count, date in query_top:
            grouped_data_sum[date.strftime(
                date_format_2)] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
              <span style='font-size: 18px'>{position}</span><br>
              <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 10px'>Count: {count}</span><br>
              <span style='font-size: 10px'>{clicks}</span> <span style='font-size: 10px; margin-left: 10px'>R: {int(impression)}</span>
              </div>"""
            total_position += position
            total_clicks += clicks
            total_impression += impression
            total_count += count
            if position > 0:
                count_for_avg += 1

        if count_for_avg > 0:
            total_position = round(total_position / count_for_avg, 2)

        grouped_data_sum["result"] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
              <span style='font-size: 18px'>{total_position}</span><br>
              <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 5px'>Count: {total_count}</span><br>
              <span style='font-size: 10px'>{total_clicks}</span> <span style='font-size: 10px; margin-left: 5px'>R: {int(total_impression)}</span>
              </div>"""

        query_front.append(grouped_data_sum)

    url_front = []
    for top in TOP:
        grouped_data_sum = {
            "query":
                f"<div style='width:355px; height: 55px; overflow: auto; white-space: nowrap;'><span>TOP {top}</span></div>"
        }
        query_top = await _get_top_url(start_date, end_date, top, async_session)
        query_top.sort(key=lambda x: x[-1])
        total_position, total_clicks, total_impression, total_count, count_for_avg = 0, 0, 0, 0, 0
        for position, clicks, impression, count, date in query_top:
            grouped_data_sum[date.strftime(
                date_format_2)] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
              <span style='font-size: 18px'>{position}</span><br>
              <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 10px'>Count: {count}</span><br>
              <span style='font-size: 10px'>{clicks}</span> <span style='font-size: 10px; margin-left: 10px'>R: {int(impression)}</span>
              </div>"""
            total_position += position
            total_clicks += clicks
            total_impression += impression
            total_count += count
            if position > 0:
                count_for_avg += 1

        if count_for_avg > 0:
            total_position = round(total_position / count_for_avg, 2)

        grouped_data_sum["result"] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
              <span style='font-size: 18px'>{total_position}</span><br>
              <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 10px'>Count: {total_count}</span><br>
              <span style='font-size: 10px'>{total_clicks}</span> <span style='font-size: 10px; margin-left: 7px'>R: {int(total_impression)}</span>
              </div>"""

        url_front.append(grouped_data_sum)

    json_data = jsonable_encoder(data)
    json_query_top = jsonable_encoder(query_front)
    json_url_top = jsonable_encoder(url_front)
    return JSONResponse({"data": json_data,
                         "query_top": json_query_top,
                         "url_top": json_url_top}
                        )


@admin_router.post("/generate_excel_indicators")
async def generate_excel_indicators(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    wb = Workbook()
    ws = wb.active
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    main_header_day_name = []
    for i in range(int(data_request["amount"])):
        date = (start_date + timedelta(days=i)).strftime(date_format_out)
        main_header_day_name.append(get_day_of_week(date))
        main_header.append(date)
    main_header = main_header[::-1]
    main_header.insert(0, "Indicators")
    main_header.insert(1, "result")
    main_header_day_name = main_header_day_name[::-1]
    main_header_day_name.insert(0, "День недели")
    main_header_day_name.insert(1, "result")
    ws.append(main_header)
    ws.append(main_header_day_name)
    main_header = []
    for i in range(int(data_request["amount"])):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    indicators = await _get_indicators_from_db(start_date, end_date, async_session)
    grouped_data = []
    if indicators:
        indicators.sort(key=lambda x: x[0])
    try:
        grouped_data = [(key, sorted(list(group), key=lambda x: x[2])) for key, group in
                        groupby(indicators, key=lambda x: x[0])]
    except TypeError as e:
        print("error")
    if len(grouped_data) == 0:
        print("empty data")
    for el in grouped_data:
        info = {}
        res = []
        value_sum, count = 0, 0
        for name, value, date in el[1]:
            if name != "TOTAL_CTR":
                info[date.strftime(date_format_out)] = [value]
            else:
                info[date.strftime(date_format_out)] = [f"{value}%"]
            value_sum += value
            count += 1

        if el[0] in ("AVG_CLICK_POSITION", "TOTAL_CTR", "AVG_SHOW_POSITION"):
            if count > 0:
                value_sum = round(value_sum / count, 2)
        res.append(el[0])
        if el[0] != "TOTAL_CTR":
            res.append(value_sum)
        else:
            res.append(f"{value_sum}%")
        for el in main_header:
            if el in info:
                res.extend(info[el])
            else:
                res.extend([0])
        ws.append(res)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(io.BytesIO(output.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment;filename='data.xlsx'"})


@admin_router.post("/generate_excel_top")
async def generate_excel_indicators(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    wb = Workbook()
    ws = wb.active
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"])):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "TOP")
    for i in range(4):
        main_header.insert(1, "result")
    ws.append(main_header)
    header = ["Position", "Click", "Impression", "Count"] * (int(data_request["amount"]) + 1)
    header.insert(0, "")
    ws.append(header)
    ws.append(["query"])
    start = 0
    main_header = []
    for i in range(int(data_request["amount"])):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("result")
    main_header = main_header[::-1]

    TOP = 3, 5, 10, 20, 30
    for top in TOP:
        res = [f"TOP {top}"]
        info = {}
        query_top = await _get_top_query(start_date, end_date, top, async_session)
        query_top.sort(key=lambda x: x[-1])
        total_position, total_clicks, total_impression, total_count, count_for_avg = 0, 0, 0, 0, 0
        for position, clicks, impression, count, date in query_top:
            info[date.strftime(date_format_out)] = [position, clicks, impression, count]
            total_position += position
            total_clicks += clicks
            total_impression += impression
            total_count += count
            if position > 0:
                count_for_avg += 1

        if count_for_avg > 0:
            total_position = round(total_position / count_for_avg, 2)
        info["result"] = [total_position, total_clicks, total_impression, total_count]
        for el in main_header:
            if el in info:
                res.extend(info[el])
            else:
                res.extend([0, 0, 0, 0])
        ws.append(res)

    ws.append(["url"])

    TOP = 3, 5, 10, 20, 30
    for top in TOP:
        res = [f"TOP {top}"]
        info = {}
        url_top = await _get_top_url(start_date, end_date, top, async_session)
        url_top.sort(key=lambda x: x[-1])
        total_position, total_clicks, total_impression, total_count, count_for_avg = 0, 0, 0, 0, 0
        for position, clicks, impression, count, date in url_top:
            info[date.strftime(date_format_out)] = [position, clicks, impression, count]
            total_position += position
            total_clicks += clicks
            total_impression += impression
            total_count += count
            if position > 0:
                count_for_avg += 1

        if count_for_avg > 0:
            total_position = round(total_position / count_for_avg, 2)
        info["result"] = [total_position, total_clicks, total_impression, total_count]
        for el in main_header:
            if el in info:
                res.extend(info[el])
            else:
                res.extend([0, 0, 0, 0])
        ws.append(res)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(io.BytesIO(output.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment;filename='data.xlsx'"})


@admin_router.post("/generate_csv_indicators/")
async def generate_excel(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    ws = []
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    main_header_day_name = []
    for i in range(int(data_request["amount"])):
        date = (start_date + timedelta(days=i)).strftime(date_format_out)
        main_header_day_name.append(get_day_of_week(date))
        main_header.append(date)
    main_header = main_header[::-1]
    main_header.insert(0, "Indicators")
    main_header_day_name = main_header_day_name[::-1]
    main_header_day_name.insert(0, "День недели")
    ws.append(main_header)
    ws.append(main_header_day_name)
    main_header = []
    for i in range(int(data_request["amount"])):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    indicators = await _get_indicators_from_db(start_date, end_date, async_session)
    grouped_data = []
    if indicators:
        indicators.sort(key=lambda x: x[0])
    try:
        grouped_data = [(key, sorted(list(group), key=lambda x: x[2])) for key, group in
                        groupby(indicators, key=lambda x: x[0])]
    except TypeError as e:
        print("error")
    if len(grouped_data) == 0:
        print("empty data")
    for el in grouped_data:
        info = {}
        res = []
        value_sum, count = 0, 0
        for name, value, date in el[1]:
            if name != "TOTAL_CTR":
                info[date.strftime(date_format_out)] = [value]
            else:
                info[date.strftime(date_format_out)] = [f"{value}%"]
            value_sum += value
            count += 1

        if el[0] in ("AVG_CLICK_POSITION", "TOTAL_CTR", "AVG_SHOW_POSITION"):
            if count > 0:
                value_sum = round(value_sum / count, 2)
        if el[0] != "TOTAL_CTR":
            res.append(value_sum)
        else:
            res.append(f"{value_sum}%")
        res.append(value_sum)
        for el in main_header:
            if el in info:
                res.extend(info[el])
            else:
                res.extend([0])
        ws.append(res)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(ws)
    output.seek(0)

    return StreamingResponse(content=output.getvalue(),
                             headers={"Content-Disposition": "attachment;filename='data.csv'"})


@admin_router.post("/generate_csv_top")
async def generate_excel_indicators(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    ws = []
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"])):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "TOP")
    for i in range(4):
        main_header.insert(1, "result")
    ws.append(main_header)
    header = ["Position", "Click", "Impression", "clicks"] * (int(data_request["amount"]) + 1)
    header.insert(0, "")
    ws.append(header)
    ws.append(["query"])
    start = 0
    main_header = []
    for i in range(int(data_request["amount"])):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("result")
    main_header = main_header[::-1]

    TOP = 3, 5, 10, 20, 30
    for top in TOP:
        res = [f"TOP {top}"]
        info = {}
        query_top = await _get_top_query(start_date, end_date, top, async_session)
        query_top.sort(key=lambda x: x[-1])
        total_position, total_clicks, total_impression, total_count, count_for_avg = 0, 0, 0, 0, 0
        for position, clicks, impression, count, date in query_top:
            info[date.strftime(date_format_out)] = [position, clicks, impression, count]
            total_position += position
            total_clicks += clicks
            total_impression += impression
            total_count += count
            if position > 0:
                count_for_avg += 1

        if count_for_avg > 0:
            total_position = round(total_position / count_for_avg, 2)
        info["result"] = [total_position, total_clicks, total_impression, total_count]
        for el in main_header:
            if el in info:
                res.extend(info[el])
            else:
                res.extend([0, 0, 0, 0])
        ws.append(res)

    ws.append(["url"])

    TOP = 3, 5, 10, 20, 30
    for top in TOP:
        res = [f"TOP {top}"]
        info = {}
        url_top = await _get_top_url(start_date, end_date, top, async_session)
        url_top.sort(key=lambda x: x[-1])
        total_position, total_clicks, total_impression, total_count, count_for_avg = 0, 0, 0, 0, 0
        for position, clicks, impression, count, date in url_top:
            info[date.strftime(date_format_out)] = [position, clicks, impression, count]
            total_position += position
            total_clicks += clicks
            total_impression += impression
            total_count += count
            if position > 0:
                count_for_avg += 1

        if count_for_avg > 0:
            total_position = round(total_position / count_for_avg, 2)
        info["result"] = [total_position, total_clicks, total_impression, total_count]
        for el in main_header:
            if el in info:
                res.extend(info[el])
            else:
                res.extend([0, 0, 0, 0])
        ws.append(res)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(ws)
    output.seek(0)

    return StreamingResponse(content=output.getvalue(),
                             headers={"Content-Disposition": "attachment;filename='data.csv'"})


@admin_router.get("/menu/merge_database/")
async def show_menu_merge_page(request: Request,
                               user: User = Depends(current_user),
                               session: AsyncSession = Depends(get_db_general)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    all_dates = await get_all_dates(async_session, QueryUrlsMergeLogs)
    group_name = request.session["group"].get("name", "")
    config_names = [elem[0] for elem in (await get_config_names(session, user, group_name))]

    group_names = await get_group_names(session, user)

    return templates.TemplateResponse("merge_database.html",
                                      {"request": request,
                                       "user": user,
                                       "config_names": config_names,
                                       "group_names": group_names,
                                       "all_dates": all_dates})


@admin_router.get("/info-merge")
async def get_info_merge(request: Request,
                         user: User = Depends(current_user),
                         session: AsyncSession = Depends(get_db_general)):
    date = request.query_params.get("date")
    group_name = request.session["group"].get("name", "")
    config_names = [elem[0] for elem in (await get_config_names(session, user, group_name))]

    group_names = await get_group_names(session, user)

    return templates.TemplateResponse("query-url-merge.html",
                                      {"request": request,
                                       "user": user,
                                       "config_names": config_names,
                                       "group_names": group_names,
                                       "date": date})


@admin_router.post("/get-merge")
async def post_info_merge(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    if data_request["sort_result"]:
        if data_request["search_text_url"] == "" and data_request["search_text_query"] == "":
            urls = await _get_merge_with_pagination_sort(data_request["date"], data_request["sort_desc"],
                                                         data_request["start"], data_request["length"],
                                                         async_session)
        else:
            urls = await _get_merge_with_pagination_and_like_sort(data_request["date"], data_request["search_text_url"],
                                                                  data_request["search_text_query"],
                                                                  data_request["sort_desc"],
                                                                  data_request["start"], data_request["length"],
                                                                  async_session)
    else:
        if data_request["search_text_url"] == "" and data_request["search_text_query"] == "":
            urls = await _get_merge_with_pagination(data_request["date"], data_request["start"], data_request["length"],
                                                    async_session)
        else:
            urls = await _get_merge_with_pagination_and_like(data_request["date"], data_request["search_text_url"],
                                                             data_request["search_text_query"],
                                                             data_request["start"], data_request["length"],
                                                             async_session)
    if not urls or len(urls) == 0:
        return JSONResponse({"data": []})
    data = []
    all_queries = list()
    print(urls)
    for el in urls:
        all_queries.extend(el[1])
    queries = await _get_merge_query(start_date, end_date, all_queries, async_session)
    if queries:
        queries.sort(key=lambda x: x[-1])
    grouped_data = dict([(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                         groupby(queries, key=lambda x: x[-1])])
    for el in urls:
        parent_clicks, parent_position, parent_impression, parent_ctr, parent_count = 0, 0, 0, 0, 0
        url, queries = el[0], el[1]
        res = {"url":
                   f"<div style='width:355px; height: 55px; overflow: auto; white-space: nowrap;'><span>{el[0]}</span></div>",
               "queries": "", "count": 0}
        for query in queries:
            query = grouped_data.get(query, None)
            if query:
                res["count"] += 1
                res[
                    "queries"] += f"<div style='width:355px; height: 55px; overflow: auto; text-align: center; white-space: nowrap;'><span>{query[0][5]}</span></div>"

                data_dict = {}
                for el in query:
                    date = el[0]  # Преобразуем дату в строку
                    values = el  # Остальные значения кортежа
                    data_dict[date] = values

                total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
                print(data_dict)
                current_date = start_date
                prev_stat = (-inf, -inf, -inf, -inf)
                while current_date <= end_date:
                    stat = data_dict.get(current_date, (-444, 0, 0, 0, 0, 0))
                    up = round(stat[1] - prev_stat[1], 2)
                    color = "#9DE8BD"
                    color_text = "green"
                    if up > 0:
                        color = "#FDC4BD"
                        color_text = "red"
                    if stat[1] <= 3:
                        color = "#B4D7ED"
                        color_text = "blue"
                    res[current_date.strftime(date_format_2)] = ''.join(
                        res.get(current_date.strftime(date_format_2), []))
                    if stat[0] != -444:
                        res[current_date.strftime(
                            date_format_2)] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: {color}'>
                                      <span style='font-size: 18px'>{stat[1]}</span><span style="margin-left: 5px; font-size: 10px; color: {color_text}">{abs(up)}</span><br>
                                      <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 20px'>CTR {stat[4]}%</span><br>
                                      <span style='font-size: 10px'>{stat[2]}</span> <span style='font-size: 10px; margin-left: 20px'>R {stat[3]}%</span>
                                      </div>"""
                    else:
                        res[current_date.strftime(
                            date_format_2)] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #B9BDBC'>
                                        <span style='font-size: 18px'><span style='color:red'>NAN</span></span><br>
                                        <span style='font-size: 10px'>Клики</span><span style='font-size: 10px; margin-left: 10px'>CTR <span style='color:red'>NAN%</span></span><br>
                                        <span style='font-size: 10px'><span style='color:red'>NAN</span></span> <span style='font-size: 10px; margin-left: 30px'>R <span style='color:red'>NAN%</span></span>
                                        </div>"""
                    total_clicks += stat[2]
                    position += stat[1]
                    impressions += stat[3]
                    ctr += stat[4]
                    if stat[1] > 0:
                        count += 1
                    prev_stat = (0, stat[1], stat[2], stat[3], stat[4])
                    if current_date == end_date:
                        res["result"] = res.get("result", "")
                        if count > 0:
                            total_position = round(position / count, 2)
                            total_ctr = round(ctr / count, 2)
                            res["result"] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
                                      <span style='font-size: 15px'>Позиция:{total_position}</span>
                                      <span style='font-size: 15px'>Клики:{total_clicks}</span>
                                      <span style='font-size: 9px'>Показы:{impressions}</span>
                                      <span style='font-size: 7px'>ctr:{total_ctr}%</span>
                                      </div>"""
                        else:
                            total_position = 0
                            total_ctr = 0
                            res["result"] += f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
                                      <span style='font-size: 15px'>Позиция:{0}</span>
                                      <span style='font-size: 15px'>Клики:{total_clicks}</span>
                                      <span style='font-size: 9px'>Показы:{impressions}</span>
                                      <span style='font-size: 9px'>ctr:{0}%</span>
                                      </div>"""
                        parent_clicks += total_clicks
                        print("sum", total_position)
                        parent_position += total_position
                        parent_impression += impressions
                        parent_ctr += total_ctr
                        if total_position > 0:
                            parent_count += 1
                    current_date += timedelta(days=1)

        if parent_count > 0:
            parent_position = round(parent_position / parent_count, 2)
            parent_ctr = round(parent_ctr / parent_count, 2)
            res["parent_result"] = f"""<div style='height: 55px; width: 100px; margin: 0px; padding: 0px; background-color: #9DE8BD'>
                                          <span style='font-size: 15px'>Позиция:{parent_position}</span>
                                          <span style='font-size: 15px'>Клики:{parent_clicks}</span>
                                          <span style='font-size: 9px'>Показы:{parent_impression}</span>
                                          <span style='font-size: 7px'>ctr:{parent_ctr}%</span>
                                          </div>"""

        data.append(res)
    json_data = jsonable_encoder(data)
    #
    # # return JSONResponse({"data": json_data, "recordsTotal": limit, "recordsFiltered": 50000})
    return JSONResponse({"data": json_data})


@admin_router.post("/generate_excel_merge/")
async def generate_excel(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    wb = Workbook()
    ws = wb.active
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "Url")
    main_header.insert(1, "Queries")
    for i in range(4):
        main_header.insert(2, "Result")
    for i in range(4):
        main_header.insert(1, "Parent Result")
    ws.append(main_header)
    header = ["Position", "Click", "R", "CTR"] * (int(data_request["amount"]) + 3)
    header.insert(0, "")
    header.insert(5, "")
    ws.append(header)
    start = 0
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("Result")
    main_header = main_header[::-1]
    while True:
        start_el = (start * 50)
        if data_request["sort_result"]:
            if data_request["search_text_url"] == "" and data_request["search_text_query"] == "":
                urls = await _get_merge_with_pagination_sort(data_request["date"], data_request["sort_desc"],
                                                             start_el, data_request["length"],
                                                             async_session)
            else:
                urls = await _get_merge_with_pagination_and_like_sort(data_request["date"],
                                                                      data_request["search_text_url"],
                                                                      data_request["search_text_query"],
                                                                      data_request["sort_desc"],
                                                                      start_el, data_request["length"],
                                                                      async_session)
        else:
            if data_request["search_text_url"] == "" and data_request["search_text_query"] == "":
                urls = await _get_merge_with_pagination(data_request["date"], start_el,
                                                        data_request["length"],
                                                        async_session)
            else:
                urls = await _get_merge_with_pagination_and_like(data_request["date"], data_request["search_text_url"],
                                                                 data_request["search_text_query"],
                                                                 start_el, data_request["length"],
                                                                 async_session)
        start += 1
        if not urls or len(urls) == 0:
            break
        all_queries = list()
        for el in urls:
            all_queries.extend(el[1])
        queries = await _get_merge_query(start_date, end_date, all_queries, async_session)
        if queries:
            queries.sort(key=lambda x: x[-1])
        grouped_data = dict([(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                             groupby(queries, key=lambda x: x[-1])])
        for url, queries in urls:
            parent_res = []
            parent_clicks, parent_position, parent_impression, parent_ctr, parent_count = 0, 0, 0, 0, 0
            for query in queries:
                res = []
                res.append(url)
                res.append(query)
                el = grouped_data.get(query, None)
                info = {}
                if el:
                    total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
                    for k, stat in enumerate(el):
                        info[stat[0].strftime(date_format_out)] = [stat[1], stat[2], stat[3], stat[4]]
                        total_clicks += stat[2]
                        position += stat[1]
                        impressions += stat[3]
                        ctr += stat[4]
                        if stat[1] > 0:
                            count += 1
                    if count > 0:
                        total_position = round(position / count, 2)
                        total_ctr = round(ctr / count, 2)
                        info["Result"] = [total_position, total_clicks, impressions, total_ctr]
                    else:
                        total_position = 0
                        total_ctr = 0
                        info["Result"] = [total_position, total_clicks, impressions, total_ctr]
                    parent_impression += impressions
                    parent_position += total_position
                    parent_clicks += total_clicks
                    parent_ctr += total_ctr
                    for el in main_header:
                        if el in info:
                            res.extend(info[el])
                        else:
                            res.extend([0, 0, 0, 0])
                parent_res.append(res)

            for parent in parent_res:
                parent_true = [parent_position, parent_clicks, parent_impression, parent_ctr]
                parent_true.extend(parent)
                ws.append(parent_true)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(io.BytesIO(output.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment;filename='data.xlsx'"})


@admin_router.post("/generate_csv_merge/")
async def generate_excel(request: Request, data_request: dict, user: User = Depends(current_user)):
    DATABASE_NAME = request.session['config'].get('database_name', "")
    group = request.session['group'].get('name', '')
    async_session = await connect_db(DATABASE_NAME, group)
    ws = []
    start_date = datetime.strptime(data_request["start_date"], date_format_2)
    end_date = datetime.strptime(data_request["end_date"], date_format_2)
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header = main_header[::-1]
    main_header.insert(0, "Url")
    main_header.insert(1, "Queries")
    for i in range(4):
        main_header.insert(2, "Result")
    for i in range(4):
        main_header.insert(1, "Parent Result")
    ws.append(main_header)
    header = ["Position", "Click", "R", "CTR"] * (int(data_request["amount"]) + 3)
    header.insert(0, "")
    header.insert(5, "")
    ws.append(header)
    start = 0
    main_header = []
    for i in range(int(data_request["amount"]) + 1):
        main_header.append((start_date + timedelta(days=i)).strftime(date_format_out))
    main_header.append("Result")
    main_header = main_header[::-1]
    while True:
        start_el = (start * 50)
        if data_request["sort_result"]:
            if data_request["search_text_url"] == "" and data_request["search_text_query"] == "":
                urls = await _get_merge_with_pagination_sort(data_request["date"], data_request["sort_desc"],
                                                             start_el, data_request["length"],
                                                             async_session)
            else:
                urls = await _get_merge_with_pagination_and_like_sort(data_request["date"],
                                                                      data_request["search_text_url"],
                                                                      data_request["search_text_query"],
                                                                      data_request["sort_desc"],
                                                                      start_el, data_request["length"],
                                                                      async_session)
        else:
            if data_request["search_text_url"] == "" and data_request["search_text_query"] == "":
                urls = await _get_merge_with_pagination(data_request["date"], start_el,
                                                        data_request["length"],
                                                        async_session)
            else:
                urls = await _get_merge_with_pagination_and_like(data_request["date"], data_request["search_text_url"],
                                                                 data_request["search_text_query"],
                                                                 start_el, data_request["length"],
                                                                 async_session)
        start += 1
        if not urls or len(urls) == 0:
            break
        all_queries = list()
        for el in urls:
            all_queries.extend(el[1])
        queries = await _get_merge_query(start_date, end_date, all_queries, async_session)
        if queries:
            queries.sort(key=lambda x: x[-1])
        grouped_data = dict([(key, sorted(list(group), key=lambda x: x[0])) for key, group in
                             groupby(queries, key=lambda x: x[-1])])
        for url, queries in urls:
            parent_res = []
            parent_clicks, parent_position, parent_impression, parent_ctr, parent_count = 0, 0, 0, 0, 0
            for query in queries:
                res = []
                res.append(url)
                res.append(query)
                el = grouped_data.get(query, None)
                info = {}
                if el:
                    total_clicks, position, impressions, ctr, count = 0, 0, 0, 0, 0
                    for k, stat in enumerate(el):
                        info[stat[0].strftime(date_format_out)] = [stat[1], stat[2], stat[3], stat[4]]
                        total_clicks += stat[2]
                        position += stat[1]
                        impressions += stat[3]
                        ctr += stat[4]
                        if stat[1] > 0:
                            count += 1
                    if count > 0:
                        total_position = round(position / count, 2)
                        total_ctr = round(ctr / count, 2)
                        info["Result"] = [total_position, total_clicks, impressions, total_ctr]
                    else:
                        total_position = 0
                        total_ctr = 0
                        info["Result"] = [total_position, total_clicks, impressions, total_ctr]
                    parent_impression += impressions
                    parent_position += total_position
                    parent_clicks += total_clicks
                    parent_ctr += total_ctr
                    for el in main_header:
                        if el in info:
                            res.extend(info[el])
                        else:
                            res.extend([0, 0, 0, 0])
                parent_res.append(res)
                print(res)
                print(parent_res)

            for parent in parent_res:
                parent_true = [parent_position, parent_clicks, parent_impression, parent_ctr]
                parent_true.extend(parent)
                ws.append(parent_true)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(ws)
    output.seek(0)

    return StreamingResponse(content=output.getvalue(),
                             headers={"Content-Disposition": "attachment;filename='data.csv'"})


@admin_router.get("/profile/{username}")
async def show_profile(request: Request,
                       username: str,
                       user=Depends(current_user),
                       session: AsyncSession = Depends(get_db_general)):
    group_name = request.session["group"].get("name", "")
    config_names = [elem[0] for elem in (await get_config_names(session, user, group_name))]

    group_names = await get_group_names(session, user)

    return templates.TemplateResponse("profile.html",
                                      {"request": request,
                                       "user": user,
                                       "config_names": config_names,
                                       "group_names": group_names})


@admin_router.get("/superuser/{username}")
async def show_superuser(
        request: Request,
        user=Depends(current_user),
        session: AsyncSession = Depends(get_db_general),
        required: bool = Depends(RoleChecker(required_permissions={"Administrator", "Superuser"}))
):
    group_name = request.session["group"].get("name", "")
    config_names = [elem[0] for elem in (await get_config_names(session, user, group_name))]

    group_names = await get_group_names(session, user)

    return templates.TemplateResponse("superuser.html",
                                      {"request": request,
                                       "user": user,
                                       "config_names": config_names,
                                       "group_names": group_names})
